"""
Module xuất báo cáo kết quả phân tích.
Hỗ trợ: CSV, Excel (có format màu sắc).
"""

import io
import pandas as pd


def results_to_dataframe(results):
    """Chuyển kết quả phân tích thành DataFrame."""
    df = pd.DataFrame([{
        'Bình luận': r['text'],
        'Kết quả': r['label'],
        'Emoji': r['emoji'],
        'Độ tin cậy': f"{r['confidence']:.1%}",
        'P(Tiêu cực)': f"{r['probs'][0]:.3f}",
        'P(Trung tính)': f"{r['probs'][1]:.3f}",
        'P(Tích cực)': f"{r['probs'][2]:.3f}",
    } for r in results])
    return df


def export_to_csv(results):
    """Xuất kết quả ra CSV bytes (để download)."""
    df = results_to_dataframe(results)
    return df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')


def export_to_excel(results, sheet_name="Kết quả phân tích"):
    """
    Xuất kết quả ra Excel bytes (có format màu sắc).
    """
    df = results_to_dataframe(results)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Auto-fit column widths
        for col_idx, column in enumerate(df.columns, 1):
            max_length = max(
                len(str(column)),
                df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) > 0 else 0
            )
            worksheet.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'AA'].width = min(max_length + 3, 50)

        # Color-code the result column
        from openpyxl.styles import PatternFill, Font, Alignment

        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        green_font = Font(color='006100')
        yellow_font = Font(color='9C6500')
        red_font = Font(color='9C0006')

        # Header styling
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data row styling
        for row_idx in range(2, len(df) + 2):
            result_cell = worksheet.cell(row=row_idx, column=2)  # "Kết quả" column
            label = result_cell.value

            if label == 'TÍCH CỰC':
                result_cell.fill = green_fill
                result_cell.font = green_font
            elif label == 'TIÊU CỰC':
                result_cell.fill = red_fill
                result_cell.font = red_font
            elif label == 'TRUNG TÍNH':
                result_cell.fill = yellow_fill
                result_cell.font = yellow_font

    output.seek(0)
    return output.getvalue()


def create_summary_text(results):
    """Tạo text tóm tắt kết quả phân tích."""
    total = len(results)
    if total == 0:
        return "Không có dữ liệu."

    pos = sum(1 for r in results if r['pred'] == 2)
    neu = sum(1 for r in results if r['pred'] == 1)
    neg = sum(1 for r in results if r['pred'] == 0)
    avg_conf = sum(r['confidence'] for r in results) / total

    summary = f"""📊 TÓM TẮT KẾT QUẢ PHÂN TÍCH
{'='*40}
📝 Tổng bình luận: {total}
✅ Tích cực: {pos} ({pos/total:.1%})
😐 Trung tính: {neu} ({neu/total:.1%})
❌ Tiêu cực: {neg} ({neg/total:.1%})
📈 Độ tin cậy trung bình: {avg_conf:.1%}
{'='*40}
"""
    return summary
